"""Geração do relatório FECHO_POS_DOP (openpyxl).

Porte de `infrastructure/report.py` do MozaOps v1. Fiel ao template do
Departamento de Meios de Pagamento e Canais (DOP). O layout segue o
ficheiro-modelo `FECHO_POS_DOP 21 a 28 de Junho-2026.xlsx`:

  · Todas as folhas de dados começam na coluna B (a coluna A fica vazia).
  · Resumo ....................... título em B3, dois blocos (validação e casos).
  · Detalhes Validacao .......... cabeçalho na linha 2, dados a partir da 3.
  · Total Casos Pendentes na SIMO cabeçalho na linha 2, só o que falta tratar.

São só estas três. O template do DOP trazia ainda o dump em bruto do Banka e a
folha «SQL» com a query de extracção — ambas saíram por não terem uso nenhum a
jusante: quem lê o relatório quer o apuramento, não a matéria-prima nem a forma
de a obter.

Todo o texto destas folhas é **conteúdo** — logo, em português.
"""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

MOZA_RED = "FFC00000"
LIGHT_GREY = "FFF2F2F2"

MONEY_FORMAT = "#,##0.00"
# Formato «contabilístico» do template para os montantes de detalhe.
ACCOUNTING_FORMAT = r'_-* #,##0.00_-;\-* #,##0.00_-;_-* "-"??_-;_-@_-'

# Diferenças: o sinal é sempre explícito, `+` incluído. Numa diferença o sinal é a
# informação — o Banka creditou a mais ou a menos são problemas opostos —, e sem o
# `+` um valor positivo lê-se como um montante qualquer. As secções do formato são
# positivo;negativo;zero(;texto): o zero fica sem sinal, e a variante contabilística
# mantém o alinhamento do template na folha de detalhe.
SIGNED_MONEY_FORMAT = "+#,##0.00;-#,##0.00;#,##0.00"
SIGNED_ACCOUNTING_FORMAT = r'_-* +#,##0.00_-;\-* #,##0.00_-;_-* "-"??_-;_-@_-'
DATE_FORMAT = "dd/mm/yyyy"
NOT_APPLICABLE = "n.a"

# As folhas de dados do template deixam a coluna A vazia e começam em B.
FIRST_COLUMN = 2

MONTHS_PT = (
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
)

VALIDATION_LABELS = {
    "match": "Crédito Confere",
    "mismatch": "Fecho creditado incorrectamente",
    "missing": "Fecho Não Creditado_aguarda tratamento da SIMO",
    "zero": "Fecho zerado (sem movimento)",
    "duplicated": "Períodos duplicados_analisar individualmente",
}

CASE_STATUS_LABELS = {
    "pending": "Fecho Não Creditado_aguarda tratamento da SIMO",
    "in_review": "Em análise",
    "resolved": "Fecho Regularizado",
}

CLOSING_TYPE_LABELS = {"D": "D", "D_PLUS_1": "D+1", "NA": NOT_APPLICABLE}

DETAILS_HEADERS = [
    "POS ID", "Comerciante", "Número de Conta", "Período POS", "POS ID vs. Período",
    "Data Fecho \nSIMO", "Nº Operaç.", "Total Fecho \nSIMO", "Descritivo Fecho",
    "Data Crédito BANKA", "TOTAL FECHO \nBANKA", "TIPO Fecho", "Validação",
    "Diferença Apurada", "e-Ticket", "Data Reg.",
]

CASES_HEADERS = [
    "POS ID", "Balcão", "Unidade Negócio", "Comerciante", "Número de Conta",
    "Período POS", "Data Fecho \nSIMO", "Nº Operaç.", "Total Fecho \nSIMO",
    "TIPO Fecho", "Descritivo Fecho", "TOTAL FECHO \nBANKA", "Validação", "Data Reg.",
]

_HEADER_FILL = PatternFill("solid", fgColor=MOZA_RED)
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)
_THIN = Side(style="thin")
_HEADER_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
_TOTAL_FILL = PatternFill("solid", fgColor=LIGHT_GREY)


def build_workbook(execution: Any, details: Iterable[Any], cases: list[Any]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _add_summary_sheet(workbook, execution, cases)
    _add_details_sheet(workbook, details, cases)
    _add_pending_cases_sheet(workbook, details, cases)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _add_summary_sheet(workbook: Workbook, execution: Any, cases: list[Any]) -> None:
    sheet = workbook.create_sheet("Resumo")
    summary = execution.summary or {}

    sheet["B3"] = _summary_title(execution.periodStart, execution.periodEnd)
    sheet["B3"].font = Font(bold=True, size=13)

    _write_header(sheet, 6, [
        "Descrição", "N° Fechos", "Montante de Fecho's Portal SIMO",
        "Montante de Fechos Creditados Banka", "Total (diferença apurada)",
    ])

    # «Fecho creditado incorrectamente» é a linha única de tudo o que não conferiu:
    # creditado a menos/mais, não creditado e períodos duplicados. Para o DOP o
    # crédito não bateu certo — a causa (diferença, ausência de crédito ou
    # duplicação) lê-se no detalhe, não no resumo.
    #
    # Do lado do Banka entram os incorrectos e os duplicados: nas chaves duplicadas
    # o Banka duplica também, e o crédito foi feito — só não se pode conferir por
    # soma. Os não creditados são os únicos sem contrapartida no Banka.
    matched_simo = _amount(summary.get("simoAmountMatched"))
    matched_banka = _amount(summary.get("bankaAmountMatched"))
    matched_count = summary.get("matched", 0)
    incorrect_simo = (
        _amount(summary.get("simoAmountMismatched"))
        + _amount(summary.get("simoAmountDuplicated"))
        + _amount(summary.get("simoAmountMissing"))
    )
    incorrect_banka = _amount(summary.get("bankaAmountMismatched")) + _amount(
        summary.get("bankaAmountDuplicated")
    )
    incorrect_count = (
        summary.get("mismatchCount", 0)
        + summary.get("duplicatedPeriods", 0)
        + summary.get("missingCount", 0)
    )

    _write_row(sheet, 7, [
        VALIDATION_LABELS["mismatch"],
        incorrect_count, incorrect_simo, incorrect_banka, incorrect_banka - incorrect_simo,
    ])
    _write_row(sheet, 8, [
        "Crédito Confere",
        matched_count, matched_simo, matched_banka, matched_banka - matched_simo,
    ])
    _write_row(sheet, 9, [
        "Total",
        incorrect_count + matched_count,
        incorrect_simo + matched_simo,
        incorrect_banka + matched_banka,
        (incorrect_banka + matched_banka) - (incorrect_simo + matched_simo),
    ], total=True)

    sheet.merge_cells("B13:D13")
    sheet["B13"] = "Total Casos Pendentes na SIMO"
    sheet["B13"].font = Font(bold=True, size=12)

    _write_header(sheet, 16, ["Descrição", "N° Fechos", "Montante de Fecho's"])

    # Espelha a folha «Total Casos Pendentes na SIMO»: o que já foi regularizado e,
    # a seguir, tudo o que continua por tratar numa só linha — não creditados,
    # creditados a menos/mais e duplicados —, com a mesma regra do bloco de cima.
    resolved = [case for case in cases if case.status == "resolved"]
    awaiting = [case for case in cases if case.status != "resolved"]
    awaiting_count = len(awaiting) + summary.get("duplicatedPeriods", 0)
    awaiting_simo = _sum_simo(awaiting) + _amount(summary.get("simoAmountDuplicated"))

    _write_row(sheet, 17, ["Fecho Regularizado", len(resolved), _sum_simo(resolved)])
    _write_row(sheet, 18, [VALIDATION_LABELS["mismatch"], awaiting_count, awaiting_simo])
    _write_row(sheet, 19, [
        "Total",
        len(resolved) + awaiting_count,
        _sum_simo(resolved) + awaiting_simo,
    ], total=True)

    _set_widths(sheet, [48, 12, 32, 34, 26])
    _set_format(sheet, ["D", "E"], MONEY_FORMAT, rows=range(7, 10))
    # F é «Total (diferença apurada)» — leva sinal.
    _set_format(sheet, ["F"], SIGNED_MONEY_FORMAT, rows=range(7, 10))
    _set_format(sheet, ["D"], MONEY_FORMAT, rows=range(17, 20))


def _add_details_sheet(workbook: Workbook, details: Iterable[Any], cases: list[Any]) -> None:
    sheet = workbook.create_sheet("Detalhes Validacao")
    _write_header(sheet, 2, DETAILS_HEADERS)

    # e-Ticket e Data Reg. são colunas do modelo, preenchidas quando a chave tem
    # um caso em tratamento — vazias nos fechos que conferem.
    case_by_key = {case.key: case for case in cases}
    row_number = 3
    for detail in details:
        case = case_by_key.get(detail.key)
        _write_row(sheet, row_number, [
            detail.posId,
            detail.merchant,
            detail.accountNumber,
            detail.period,
            detail.key,
            detail.simoClosingDate,
            detail.operationNumber,
            detail.simoClosingTotal,
            detail.closingDescription or NOT_APPLICABLE,
            detail.bankaCreditDate or NOT_APPLICABLE,
            detail.bankaClosingTotal if detail.bankaClosingTotal is not None else 0,
            CLOSING_TYPE_LABELS.get(detail.closingType, NOT_APPLICABLE),
            VALIDATION_LABELS.get(detail.validation, detail.validation),
            detail.difference if detail.difference is not None else NOT_APPLICABLE,
            (case.eTicket if case else None) or "",
            (case.resolvedAt if case else None) or "",
        ])
        row_number += 1

    _set_widths(sheet, [10, 32, 16, 12, 18, 14, 10, 18, 34, 16, 18, 10, 38, 16, 12, 12])
    # I é «Total Fecho SIMO»; O é «Diferença Apurada» — só esta leva sinal.
    _set_format(sheet, ["I"], ACCOUNTING_FORMAT, rows=range(3, row_number))
    _set_format(sheet, ["O"], SIGNED_ACCOUNTING_FORMAT, rows=range(3, row_number))
    _set_format(sheet, ["L"], MONEY_FORMAT, rows=range(3, row_number))
    _set_format(sheet, ["G", "K", "Q"], DATE_FORMAT, rows=range(3, row_number))
    sheet.freeze_panes = "A3"


def _add_pending_cases_sheet(workbook: Workbook, details: Iterable[Any], cases: list[Any]) -> None:
    """A lista do que fica por tratar — o que se leva à SIMO.

    Entram os três problemas que exigem acção, e só enquanto não estiverem
    tratados: creditado incorrectamente, não creditado e períodos duplicados. Um
    caso regularizado sai daqui (o Resumo é que o contabiliza); os duplicados não
    geram caso nenhum, por isso vêm dos detalhes, um por fecho — é fecho a fecho
    que se desfaz a duplicação.

    Os duplicados vão rotulados como «Fecho creditado incorrectamente»: para o DOP
    o crédito não bateu certo, e a duplicação do período é a causa, não uma
    categoria à parte. O Resumo soma-os na mesma linha, para as duas folhas
    fecharem uma com a outra — incluindo o que o Banka já creditou nessas chaves,
    onde os movimentos vêm duplicados tal como os fechos da SIMO.
    """
    sheet = workbook.create_sheet("Total Casos Pendentes na SIMO")
    _write_header(sheet, 2, CASES_HEADERS)

    detail_by_key: dict[str, Any] = {}
    for detail in details:
        detail_by_key.setdefault(detail.key, detail)

    row_number = 3
    # Incorrectos primeiro, depois não creditados: dinheiro errado antes de dinheiro
    # em falta. Um caso regularizado já não está pendente na SIMO.
    for kind in ("mismatch", "missing"):
        for case in (c for c in cases if c.type == kind and c.status != "resolved"):
            detail = detail_by_key.get(case.key)
            _write_row(sheet, row_number, [
                case.posId,
                NOT_APPLICABLE,  # Balcão (não vem nos ficheiros de entrada)
                NOT_APPLICABLE,  # Unidade Negócio (idem)
                case.merchant,
                case.accountNumber,
                case.period,
                detail.simoClosingDate if detail else NOT_APPLICABLE,
                detail.operationNumber if detail else NOT_APPLICABLE,
                case.simoAmount,
                CLOSING_TYPE_LABELS.get(detail.closingType, NOT_APPLICABLE) if detail else NOT_APPLICABLE,
                (detail.closingDescription if detail else None) or NOT_APPLICABLE,
                case.bankaAmount,
                VALIDATION_LABELS[kind],
                case.resolvedAt or NOT_APPLICABLE,
            ])
            row_number += 1

    # Duplicados: um por fecho, com o valor do próprio fecho. O crédito do Banka é
    # da chave inteira (lá os movimentos também vêm duplicados), por isso vai só na
    # primeira linha de cada chave — repeti-lo em todas inflacionaria a coluna, que
    # é exactamente o erro do VLOOKUP manual que esta automação veio corrigir.
    # Assim a coluna M soma para o mesmo que o Resumo.
    credited_keys: set[str] = set()
    for detail in (d for d in details if d.validation == "duplicated"):
        first_of_key = detail.key not in credited_keys
        credited_keys.add(detail.key)
        _write_row(sheet, row_number, [
            detail.posId,
            NOT_APPLICABLE,
            NOT_APPLICABLE,
            detail.merchant,
            detail.accountNumber,
            detail.period,
            detail.simoClosingDate,
            detail.operationNumber,
            detail.simoClosingTotal,
            CLOSING_TYPE_LABELS.get(detail.closingType, NOT_APPLICABLE),
            detail.closingDescription or NOT_APPLICABLE,
            detail.bankaClosingTotal if first_of_key and detail.bankaClosingTotal is not None
            else NOT_APPLICABLE,
            VALIDATION_LABELS["mismatch"],
            NOT_APPLICABLE,
        ])
        row_number += 1

    # Colunas (com a A vazia): G=Período POS, H=Data Fecho, J=Total SIMO,
    # M=Total Banka, O=Data Reg. — não deslocar os formatos por causa da A.
    _set_widths(sheet, [10, 8, 18, 32, 16, 12, 14, 10, 18, 10, 34, 18, 42, 12])
    _set_format(sheet, ["J", "M"], MONEY_FORMAT, rows=range(3, row_number))
    _set_format(sheet, ["H", "O"], DATE_FORMAT, rows=range(3, row_number))
    sheet.freeze_panes = "A3"


def _summary_title(start: Any, end: Any) -> str:
    start = _excel_value(start)
    end = _excel_value(end)
    if start.month == end.month:
        span = f"{start.day}  a  {end.day} de {MONTHS_PT[end.month - 1]}  {end.year}"
    else:
        span = (
            f"{start.day} de {MONTHS_PT[start.month - 1]}  a  "
            f"{end.day} de {MONTHS_PT[end.month - 1]}  {end.year}"
        )
    return f"Validação de crédito de Fechos ({span})"


def _write_header(sheet: Worksheet, row: int, values: list[str]) -> None:
    for offset, value in enumerate(values):
        cell = sheet.cell(row=row, column=FIRST_COLUMN + offset, value=value)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER


def _write_row(sheet: Worksheet, row: int, values: list[Any], *, total: bool = False) -> None:
    for offset, value in enumerate(values):
        cell = sheet.cell(row=row, column=FIRST_COLUMN + offset, value=_excel_value(value))
        if total:
            cell.font = Font(bold=True)
            cell.fill = _TOTAL_FILL


def _excel_value(value: Any) -> Any:
    """O Excel não aceita datas com fuso horário e a BD pode devolver `datetime`.

    As colunas de data guardam só a data (`Date`), logo só a data interessa —
    descarta-se a hora se por acaso vier um `datetime`.
    """
    if isinstance(value, datetime):
        return value.date()
    return value


def _set_widths(sheet: Worksheet, widths: list[int]) -> None:
    sheet.column_dimensions["A"].width = 2.5
    for offset, width in enumerate(widths):
        sheet.column_dimensions[get_column_letter(FIRST_COLUMN + offset)].width = width


def _set_format(sheet: Worksheet, columns: list[str], number_format: str, rows: range) -> None:
    for column in columns:
        for row in rows:
            sheet[f"{column}{row}"].number_format = number_format


def _sum_simo(cases: list[Any]) -> Decimal:
    return sum((case.simoAmount for case in cases), Decimal(0))


def _amount(value: Any) -> Decimal:
    return Decimal(str(value or 0))
