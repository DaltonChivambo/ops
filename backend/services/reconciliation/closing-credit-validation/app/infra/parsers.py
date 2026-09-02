"""Adaptador de leitura dos três ficheiros de entrada (openpyxl).

Porte de `infrastructure/parsers.py` do MozaOps v1. Camada de infraestrutura:
é a única parte do serviço que sabe que os dados vêm de Excel. Produz as
estruturas de `domain/models.py` e nada mais.

Cada ficheiro é validado pelos cabeçalhos esperados ANTES de ser parseado — é
isso que também apanha um ficheiro carregado no campo errado.

As posições das colunas variam entre exports, por isso a Lista de POS e o Banka
resolvem as colunas pelo NOME do cabeçalho (não por índice fixo); a SIMO mantém
índices fixos. Estruturas (verificadas nos ficheiros reais):
  Lista POS ...... folha `Export`; colunas Merchant Id · POS Id · Nome Comerciante
                   · Nº Conta · Fecho Realtime (Sim/Não), em ordem/posição variável
  Fechos SIMO .... folha 1; B Id Comerciante · C POS Id · D Período POS
                   · E Data Fecho (serial) · F Nº Operaç. · G Total Fecho (pt)
  Créditos Banka . folha `FECHO_POS`; DATA_SISTEMA · DESCRITIVO_MOV ·
                   VALOR_TRANSACAO. As posições das colunas VARIAM entre exports
                   do MIS, por isso são resolvidas pelo NOME do cabeçalho e não
                   por índice fixo (ver `parse_banka_credits`). A chave «POS ID
                   vs. Período» é sempre derivada do DESCRITIVO_MOV.
"""
from itertools import chain, islice
from typing import IO, Any, Iterator

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..domain.keys import key_from_description, normalize_pos_id
from ..domain.models import BankaCredit, BankaMovement, PosInfo, SimoClosing
from ..errors import BusinessError
from .excel import cell_date, cell_text, find_header_row, parse_number, validate_headers

SLOT_LABELS = {
    "posList": "Lista de POS",
    "simoClosings": "Fechos SIMO",
    "bankaCredits": "Créditos Banka",
}

# Índices 0-based das colunas usadas na Lista de POS e nos Fechos SIMO (estáveis).
# O Banka resolve as colunas pelo nome do cabeçalho — ver `parse_banka_credits`.
POS_LIST_COLUMNS = {"posId": 1, "merchant": 2, "accountNumber": 3, "realtime": 4}
SIMO_COLUMNS = {"posId": 2, "period": 3, "closingDate": 4, "operationNumber": 5, "total": 6}

HEADER_SEARCH_ROWS = 10
BANKA_SAMPLE_ROWS = 20  # linhas espreitadas para escolher a coluna DESCRITIVO_MOV certa


def _structure_error(slot: str, filename: str, missing: list[str]) -> BusinessError:
    return BusinessError(
        f"Dados incompletos ou em formato inválido: o ficheiro «{filename}» no campo "
        f"«{SLOT_LABELS[slot]}» não tem as colunas esperadas ({', '.join(missing)}). "
        "Verifique se carregou o ficheiro correcto e volte a submeter."
    )


def _open_sheet(stream: IO[bytes], slot: str, filename: str, sheet_name: str | None) -> Worksheet:
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 — ficheiro corrompido ou não-Excel
        raise BusinessError(
            f"Dados incompletos ou em formato inválido: não foi possível ler o ficheiro "
            f"«{filename}» no campo «{SLOT_LABELS[slot]}». Confirme que é um Excel (.xlsx) válido."
        ) from error

    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    if not workbook.sheetnames:
        raise _structure_error(slot, filename, [sheet_name or "folha de dados"])
    return workbook[workbook.sheetnames[0]]


def _header_and_rows(
    sheet: Worksheet,
    slot: str,
    filename: str,
    anchor: str,
    expected: list[str],
) -> tuple[tuple[Any, ...], Iterator[tuple[Any, ...]]]:
    """Valida os cabeçalhos e devolve `(linha de cabeçalho, iterador de dados)`.

    A folha é percorrida uma única vez (`read_only`): só as primeiras linhas ficam
    em memória, para localizar o cabeçalho; o resto é consumido em streaming.
    """
    rows = sheet.iter_rows(values_only=True)
    head: list[tuple[Any, ...]] = []
    for row in rows:
        head.append(row)
        if len(head) >= HEADER_SEARCH_ROWS:
            break

    header_index = find_header_row(head, anchor)
    if header_index is None:
        raise _structure_error(slot, filename, [anchor])

    header = head[header_index]
    missing = validate_headers(header, expected)
    if missing:
        raise _structure_error(slot, filename, missing)

    def _iter() -> Iterator[tuple[Any, ...]]:
        yield from head[header_index + 1:]
        yield from rows

    return header, _iter()


def _data_rows(
    sheet: Worksheet,
    slot: str,
    filename: str,
    anchor: str,
    expected: list[str],
) -> Iterator[tuple[Any, ...]]:
    """Como `_header_and_rows`, mas devolve só as linhas de dados."""
    _header, rows = _header_and_rows(sheet, slot, filename, anchor, expected)
    return rows


def _value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _column_index(header: tuple[Any, ...], name: str) -> int | None:
    """Índice da 1ª coluna cujo cabeçalho começa por `name` (case-insensitive)."""
    target = name.lower()
    for index, cell in enumerate(header):
        if cell_text(cell).lower().startswith(target):
            return index
    return None


def _description_column(header: tuple[Any, ...], sample: list[tuple[Any, ...]]) -> int | None:
    """Escolhe a coluna DESCRITIVO_MOV que traz a chave.

    O export do MIS pode repetir o cabeçalho `DESCRITIVO_MOV` (uma versão truncada
    e a versão completa que contém «... - NNN»). Escolhe-se a coluna cujas linhas
    de amostra produzem uma chave válida; se nenhuma o fizer, a última (a completa
    costuma vir à direita).
    """
    candidates = [
        index
        for index, cell in enumerate(header)
        if cell_text(cell).lower().startswith("descritivo")
    ]
    if len(candidates) <= 1:
        return candidates[0] if candidates else None
    for index in candidates:
        if any(key_from_description(cell_text(_value(row, index))) for row in sample):
            return index
    return candidates[-1]


def parse_pos_list(stream: IO[bytes], filename: str) -> dict[str, PosInfo]:
    """Lista de POS → `{posId: PosInfo}`. Fecho Realtime: Sim → D, Não → D+1.

    As colunas variam entre exports (o `ListaPOS_TOTAL` tem 48 colunas noutra
    ordem), por isso são resolvidas pelo NOME do cabeçalho. O POS Id é
    normalizado (sem zeros à esquerda) para casar com a SIMO e o Banka.
    """
    sheet = _open_sheet(stream, "posList", filename, "Export")
    header, rows = _header_and_rows(
        sheet,
        "posList",
        filename,
        "merchant id",
        ["merchant id", "pos id", "nome comerciante", "nº conta", "fecho realtime"],
    )

    pos_id_col = _column_index(header, "pos id")
    merchant_col = _column_index(header, "nome comerciante")
    account_col = _column_index(header, "nº conta")
    realtime_col = _column_index(header, "fecho realtime")

    result: dict[str, PosInfo] = {}
    for row in rows:
        pos_id = normalize_pos_id(cell_text(_value(row, pos_id_col)))
        if not pos_id:
            continue
        realtime = cell_text(_value(row, realtime_col)).lower()
        result[pos_id] = PosInfo(
            merchant=cell_text(_value(row, merchant_col)),
            accountNumber=cell_text(_value(row, account_col)),
            closingType=(
                "D" if realtime == "sim" else "D_PLUS_1" if realtime in ("não", "nao") else "NA"
            ),
        )
    return result


def parse_simo_closings(stream: IO[bytes], filename: str) -> list[SimoClosing]:
    """Fechos do Portal SIMO → lista de fechos válidos."""
    sheet = _open_sheet(stream, "simoClosings", filename, None)
    rows = _data_rows(
        sheet,
        "simoClosings",
        filename,
        "id comerciante",
        ["id comerciante", "pos id", "período pos", "data fecho", "total fecho"],
    )

    closings: list[SimoClosing] = []
    for row in rows:
        pos_id = normalize_pos_id(cell_text(_value(row, SIMO_COLUMNS["posId"])))
        period = parse_number(_value(row, SIMO_COLUMNS["period"]))
        total = parse_number(_value(row, SIMO_COLUMNS["total"]))
        closing_date = cell_date(_value(row, SIMO_COLUMNS["closingDate"]))
        if not pos_id or period is None or total is None or closing_date is None:
            continue
        operation = parse_number(_value(row, SIMO_COLUMNS["operationNumber"]))
        closings.append(
            SimoClosing(
                posId=pos_id,
                period=int(period),
                closingDate=closing_date,
                operationNumber=int(operation) if operation is not None else 0,
                total=total,
            )
        )
    return closings


def parse_banka_credits(stream: IO[bytes], filename: str) -> dict[str, BankaCredit]:
    """Créditos do Banka (MIS) agregados por chave «POS ID vs. Período».

    A chave é sempre derivada do DESCRITIVO_MOV — a coluna «POS ID vs. Período»
    não é usada (é considerada não-fiável e nem sempre existe no export). As
    posições das colunas variam entre exports, por isso são resolvidas pelo NOME
    do cabeçalho: DATA_SISTEMA, DESCRITIVO_MOV e VALOR_TRANSACAO.
    """
    sheet = _open_sheet(stream, "bankaCredits", filename, "FECHO_POS")
    header, rows = _header_and_rows(
        sheet,
        "bankaCredits",
        filename,
        "data_sistema",
        ["data_sistema", "descritivo", "valor_transacao"],
    )

    date_col = _column_index(header, "data_sistema")
    amount_col = _column_index(header, "valor_transacao")
    # DESCRITIVO_MOV pode aparecer duplicado; escolhe-se pela amostra a que traz a chave.
    rows = iter(rows)
    sample = list(islice(rows, BANKA_SAMPLE_ROWS))
    desc_col = _description_column(header, sample)
    rows = chain(sample, rows)

    credits: dict[str, BankaCredit] = {}
    for row in rows:
        description = cell_text(_value(row, desc_col)) if desc_col is not None else ""
        key = key_from_description(description) if description else ""
        amount = parse_number(_value(row, amount_col)) if amount_col is not None else None
        if not key or amount is None:
            continue

        credit_date = cell_date(_value(row, date_col)) if date_col is not None else None
        movement = BankaMovement(date=credit_date, amount=amount, description=description or None)
        existing = credits.get(key)
        if existing:
            existing.amount += amount
            existing.movements.append(movement)
            # Uma chave pode ter movimentos em dias diferentes; a data de crédito
            # é a do primeiro movimento, não a da primeira linha que aparece no
            # ficheiro — o export do MIS não vem por ordem cronológica.
            if credit_date and (existing.creditDate is None or credit_date < existing.creditDate):
                existing.creditDate = credit_date
        else:
            credits[key] = BankaCredit(
                amount=amount,
                creditDate=credit_date,
                description=description or None,
                movements=[movement],
            )
    return credits
